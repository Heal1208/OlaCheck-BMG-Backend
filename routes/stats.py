from flask import Blueprint, request, jsonify, send_file
from database import get_db
from utils import token_required, role_required, MANAGER_AND_ABOVE

import io
from datetime import datetime

stats_bp = Blueprint("stats", __name__, url_prefix="/api/stats")


# ================================================================
# A-11: Staff Schedule / Work Monitoring
# ================================================================
@stats_bp.route("/staff-schedule", methods=["GET"])
@token_required
@role_required(*MANAGER_AND_ABOVE)
def staff_schedule(current_user):
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_db()

    staff_list = conn.execute("""
        SELECT u.user_id, u.full_name, u.email, u.phone,
               COUNT(DISTINCT s.store_id) AS total_assigned_stores
        FROM users u
        JOIN roles r ON u.role_id = r.role_id
        LEFT JOIN stores s ON s.assigned_staff_id = u.user_id AND s.is_active = 1
        WHERE r.role_name = 'Staff' AND u.is_active = 1
        GROUP BY u.user_id
    """).fetchall()

    result = []
    for staff in staff_list:
        uid = staff["user_id"]
        checkins = conn.execute("""
            SELECT c.check_id, c.store_id, c.check_time, c.status, c.note,
                   s.store_name, s.district, s.city
            FROM store_checks c
            JOIN stores s ON c.store_id = s.store_id
            WHERE c.staff_id = ? AND DATE(c.check_time) = ?
            ORDER BY c.check_time ASC
        """, (uid, date)).fetchall()

        checkin_details = []
        for c in checkins:
            stock_count = conn.execute(
                "SELECT COUNT(*) FROM stock_entries WHERE check_id = ?", (c["check_id"],)
            ).fetchone()[0]
            expiry_count = conn.execute("""
                SELECT COUNT(*) FROM expiry_records er
                JOIN stock_entries se ON er.entry_id = se.entry_id
                WHERE se.check_id = ?
            """, (c["check_id"],)).fetchone()[0]
            checkin_details.append({
                "check_id":       c["check_id"],
                "store_id":       c["store_id"],
                "store_name":     c["store_name"],
                "district":       c["district"],
                "city":           c["city"],
                "check_time":     c["check_time"],
                "status":         c["status"],
                "note":           c["note"],
                "stock_entered":  stock_count > 0,
                "expiry_checked": expiry_count > 0,
                "steps_done":     1 + (1 if stock_count > 0 else 0) + (1 if expiry_count > 0 else 0),
            })

        total = staff["total_assigned_stores"]
        done  = len([c for c in checkin_details if c["status"] == "completed"])
        result.append({
            "user_id":        uid,
            "full_name":      staff["full_name"],
            "email":          staff["email"],
            "phone":          staff["phone"],
            "total_stores":   total,
            "done_stores":    done,
            "completion_pct": round((done / total * 100) if total > 0 else 0),
            "checkins":       checkin_details,
        })

    conn.close()
    return jsonify({"success": True, "data": {"date": date, "schedule": result}}), 200


# ================================================================
# A-12: Inventory Statistics
# ================================================================
@stats_bp.route("/inventory", methods=["GET"])
@token_required
@role_required(*MANAGER_AND_ABOVE)
def inventory_stats(current_user):
    date_from  = request.args.get("date_from", "")
    date_to    = request.args.get("date_to", "")
    store_id   = request.args.get("store_id", "")
    product_id = request.args.get("product_id", "")

    where, params = "WHERE 1=1", []
    if date_from:  where += " AND DATE(c.check_time) >= ?"; params.append(date_from)
    if date_to:    where += " AND DATE(c.check_time) <= ?"; params.append(date_to)
    if store_id:   where += " AND c.store_id = ?";          params.append(store_id)
    if product_id: where += " AND se.product_id = ?";       params.append(product_id)

    conn = get_db()

    by_product = conn.execute(f"""
        SELECT p.product_id, p.product_name, p.sku, p.unit,
               COALESCE(SUM(se.quantity_on_shelf), 0)  AS total_qty,
               COUNT(DISTINCT c.store_id)               AS store_count,
               COALESCE(AVG(se.quantity_on_shelf), 0)  AS avg_qty
        FROM stock_entries se
        JOIN store_checks c ON se.check_id  = c.check_id
        JOIN products p     ON se.product_id = p.product_id
        {where}
        GROUP BY p.product_id ORDER BY total_qty DESC
    """, params).fetchall()

    by_store = conn.execute(f"""
        SELECT s.store_id, s.store_name, s.store_type, s.district, s.city,
               COALESCE(SUM(se.quantity_on_shelf), 0) AS total_qty,
               MAX(c.check_time)                       AS last_checkin
        FROM stock_entries se
        JOIN store_checks c ON se.check_id = c.check_id
        JOIN stores s       ON c.store_id  = s.store_id
        {where}
        GROUP BY s.store_id ORDER BY total_qty DESC
    """, params).fetchall()

    timeline = conn.execute(f"""
        SELECT DATE(c.check_time)                     AS report_date,
               COALESCE(SUM(se.quantity_on_shelf), 0) AS total_qty
        FROM stock_entries se
        JOIN store_checks c ON se.check_id = c.check_id
        {where}
        GROUP BY DATE(c.check_time) ORDER BY report_date ASC LIMIT 30
    """, params).fetchall()

    by_type = conn.execute(f"""
        SELECT s.store_type,
               COALESCE(SUM(se.quantity_on_shelf), 0) AS total_qty,
               COUNT(DISTINCT s.store_id)              AS store_count
        FROM stock_entries se
        JOIN store_checks c ON se.check_id = c.check_id
        JOIN stores s       ON c.store_id  = s.store_id
        {where}
        GROUP BY s.store_type
    """, params).fetchall()

    products    = conn.execute("SELECT product_id, product_name, sku FROM products WHERE is_active=1 ORDER BY product_name").fetchall()
    stores      = conn.execute("SELECT store_id, store_name, district FROM stores WHERE is_active=1 ORDER BY store_name").fetchall()
    open_alerts = conn.execute("SELECT COUNT(*) FROM stock_alerts WHERE is_resolved=0 AND alert_type='low_stock'").fetchone()[0]
    conn.close()

    return jsonify({"success": True, "data": {
        "summary": {
            "total_products":  len(by_product),
            "total_stores":    len(by_store),
            "open_low_stock":  open_alerts,
            "grand_total_qty": sum(r["total_qty"] or 0 for r in by_product),
        },
        "by_product": [dict(r) for r in by_product],
        "by_store":   [dict(r) for r in by_store],
        "by_type":    [dict(r) for r in by_type],
        "timeline":   [dict(r) for r in timeline],
        "filter_options": {
            "products": [dict(p) for p in products],
            "stores":   [dict(s) for s in stores],
        },
    }}), 200


# ================================================================
# A-15: Export XLSX (Excel) — 2 sheets, định dạng màu sắc
# ================================================================
@stats_bp.route("/export", methods=["GET"])
@token_required
@role_required(*MANAGER_AND_ABOVE)
def export_report(current_user):
    date_from  = request.args.get("date_from", "")
    date_to    = request.args.get("date_to", "")
    store_id   = request.args.get("store_id", "")
    product_id = request.args.get("product_id", "")

    where, params = "WHERE 1=1", []
    if date_from:  where += " AND DATE(c.check_time) >= ?"; params.append(date_from)
    if date_to:    where += " AND DATE(c.check_time) <= ?"; params.append(date_to)
    if store_id:   where += " AND c.store_id = ?";          params.append(store_id)
    if product_id: where += " AND se.product_id = ?";       params.append(product_id)

    conn = get_db()
    rows = conn.execute(f"""
        SELECT s.store_name, s.district, s.city, s.store_type,
               p.product_name, p.sku,
               se.quantity_on_shelf,
               DATE(c.check_time) AS report_date,
               c.check_time       AS check_time_full,
               u.full_name        AS staff_name
        FROM stock_entries se
        JOIN store_checks c ON se.check_id   = c.check_id
        JOIN stores s       ON c.store_id    = s.store_id
        JOIN products p     ON se.product_id = p.product_id
        JOIN users u        ON c.staff_id    = u.user_id
        {where}
        ORDER BY c.check_time DESC LIMIT 5000
    """, params).fetchall()

    by_product = conn.execute(f"""
        SELECT p.product_name, p.sku,
               COALESCE(SUM(se.quantity_on_shelf), 0) AS total_qty,
               COUNT(DISTINCT c.store_id)              AS store_count
        FROM stock_entries se
        JOIN store_checks c ON se.check_id   = c.check_id
        JOIN products p     ON se.product_id = p.product_id
        {where}
        GROUP BY p.product_id ORDER BY total_qty DESC
    """, params).fetchall()

    conn.close()

    # ── Build Excel ───────────────────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"success": False, "message": "Thiếu thư viện openpyxl. Chạy: pip install openpyxl"}), 500

    wb = Workbook()

    # ── Palette ───────────────────────────────────────────────
    C_DARK   = "1A2B3C"
    C_GOLD   = "C9A227"
    C_ALT    = "F4F7FB"
    C_WHITE  = "FFFFFF"
    C_GREEN  = "1A6B3C"
    C_RED    = "C0392B"
    C_BORDER = "D1D9E6"

    def border(color=C_BORDER):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_cell(ws, row, col, value, width=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=C_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border()
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def data_cell(ws, row, col, value, alt=False, align="left", bold=False, color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name="Calibri", size=10, bold=bold, color=color or C_DARK)
        c.fill      = PatternFill("solid", fgColor=C_ALT if alt else C_WHITE)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = border()
        return c

    # ════════════════════════════════════════════════════════
    # Sheet 1: Chi tiết
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Chi tiết tồn kho"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:J1")
    t = ws1["A1"]
    t.value     = "BÁO CÁO TỒN KHO – BMG TRADING · OLASUN SUNFLOWER OIL"
    t.font      = Font(name="Calibri", bold=True, size=15, color=C_WHITE)
    t.fill      = PatternFill("solid", fgColor=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    # Sub-title
    ws1.merge_cells("A2:J2")
    s = ws1["A2"]
    s.value     = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}    |    Tổng bản ghi: {len(rows)}"
    s.font      = Font(name="Calibri", italic=True, size=10, color="5A6A7E")
    s.fill      = PatternFill("solid", fgColor="EBF0F8")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 6  # spacer

    # Headers
    HEADERS = [
        ("Cửa hàng", 26), ("Quận/Huyện", 16), ("Thành phố", 14),
        ("Loại", 14), ("Sản phẩm", 28), ("SKU", 14),
        ("Số lượng tồn", 16), ("Ngày báo cáo", 16),
        ("Thời gian check-in", 22), ("Nhân viên", 22),
    ]
    for col_i, (h, w) in enumerate(HEADERS, 1):
        hdr_cell(ws1, 4, col_i, h, w)
    ws1.row_dimensions[4].height = 28

    # Data
    for ri, row in enumerate(rows, 5):
        alt = ri % 2 == 0
        qty = row["quantity_on_shelf"] or 0
        vals = [
            row["store_name"], row["district"], row["city"], row["store_type"],
            row["product_name"], row["sku"], qty,
            row["report_date"], row["check_time_full"], row["staff_name"],
        ]
        for ci, v in enumerate(vals, 1):
            kw = dict(alt=alt)
            if ci == 7:   # Qty
                kw.update(align="center", bold=True, color=C_RED if qty < 10 else C_GREEN)
            elif ci in (1, 5):
                kw["align"] = "left"
            else:
                kw["align"] = "center"
            data_cell(ws1, ri, ci, v, **kw)
        ws1.row_dimensions[ri].height = 18

    ws1.freeze_panes = "A5"

    # ════════════════════════════════════════════════════════
    # Sheet 2: Tổng hợp sản phẩm
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Tổng hợp sản phẩm")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    t2 = ws2["A1"]
    t2.value     = "TỔNG HỢP TỒN KHO THEO SẢN PHẨM"
    t2.font      = Font(name="Calibri", bold=True, size=13, color=C_WHITE)
    t2.fill      = PatternFill("solid", fgColor=C_DARK)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    H2 = [("Tên sản phẩm", 30), ("SKU", 16), ("Tổng tồn kho", 18),
          ("Số cửa hàng", 18), ("% tổng kho", 14)]
    for ci, (h, w) in enumerate(H2, 1):
        hdr_cell(ws2, 2, ci, h, w)
    ws2.row_dimensions[2].height = 26

    grand = sum(r["total_qty"] or 0 for r in by_product) or 1
    for ri, row in enumerate(by_product, 3):
        alt = ri % 2 == 0
        qty = row["total_qty"] or 0
        pct = f"{round(qty / grand * 100, 1)}%"
        for ci, v in enumerate([row["product_name"], row["sku"], qty, row["store_count"], pct], 1):
            kw = dict(alt=alt, align="center")
            if ci == 1:  kw["align"] = "left"
            if ci == 3:  kw.update(bold=True, color=C_GREEN)
            data_cell(ws2, ri, ci, v, **kw)
        ws2.row_dimensions[ri].height = 20

    ws2.freeze_panes = "A3"

    # ── Stream ───────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"BMG_inventory_{ts}.xlsx"
    )